import openpyxl
from django.http import FileResponse, HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from reportlab.pdfgen import canvas

from calculator_projects.apps.projects.models import ProjectFact, ProjectPlan
from calculator_projects.apps.projects.utils import excel_generation_plan_fact_compare
from calculator_projects.apps.stages.models import StagePlan
from calculator_projects.apps.users.models import UserRoleTypes


def generate_pdf(request, pk):
    stage_plan = StagePlan.objects.filter(projectPlan=pk, deleted_status=False).order_by("stage_number")
    projectplan = ProjectPlan.objects.get(id=pk)

    response = FileResponse(
        generate_pdf_file(stage_plan, projectplan), as_attachment=True, filename="book_catalog.pdf"
    )
    return response


def generate_pdf_file(stage_list, project):
    from io import BytesIO

    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    p.drawString(100, 750, "Book Catalog")

    y = 700
    for stage in stage_list:
        p.drawString(100, y, f"Title: {stage.description}")
        p.drawString(100, y - 20, f"Author: {stage.start_time}")
        p.drawString(100, y - 40, f"Year: {stage.finish_time}")
        y -= 60

    p.showPage()
    p.save()

    buffer.seek(0)

    return buffer


def export_excel(request, pk):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="project.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write header row
    header = ["Паспорт проекта"]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    project = ProjectPlan.objects.get(id=pk)
    worksheet.title = str(project.id)

    worksheet.cell(row=1, column=1).value = 1
    worksheet.cell(row=1, column=2).value = "ID"
    worksheet.cell(row=1, column=3).value = str(project.id)
    worksheet.cell(row=2, column=1).value = 2
    worksheet.cell(row=2, column=2).value = "Наименование проекта"
    worksheet.cell(row=2, column=3).value = project.name
    worksheet.cell(row=3, column=1).value = 3
    worksheet.cell(row=3, column=2).value = "Заказчик"
    worksheet.cell(row=3, column=3).value = project.customer
    worksheet.cell(row=4, column=1).value = 4
    worksheet.cell(row=4, column=2).value = "Ответственные организации за реализацию проекта"
    worksheet.cell(row=4, column=3).value = project.responsible_subject_for_project
    worksheet.cell(row=5, column=1).value = 5
    worksheet.cell(row=5, column=2).value = "Основания для реализации проекта"
    worksheet.cell(row=5, column=3).value = project.legal_basis
    worksheet.cell(row=6, column=1).value = 6
    worksheet.cell(row=6, column=2).value = "Источники финансирования"
    worksheet.cell(row=6, column=3).value = project.source_of_financing
    worksheet.cell(row=7, column=1).value = 7
    worksheet.cell(
        row=7, column=2
    ).value = "Включен ли проект в государственную программу развития Республики Узбекистан"
    worksheet.cell(row=7, column=3).value = project.involved_in_development_country
    worksheet.cell(row=8, column=1).value = 8
    worksheet.cell(row=8, column=2).value = "Тип проекта"
    worksheet.cell(row=8, column=3).value = project.customer_status
    worksheet.cell(row=9, column=1).value = 9
    worksheet.cell(row=9, column=2).value = "Цель проекта"
    worksheet.cell(row=9, column=3).value = project.purpose
    worksheet.cell(row=10, column=1).value = 10
    worksheet.cell(row=10, column=2).value = "Задачи проекта"
    worksheet.cell(row=10, column=3).value = project.objective
    worksheet.cell(row=11, column=1).value = 11
    worksheet.cell(row=11, column=2).value = "Ориентировочная дата начала проекта"
    worksheet.cell(row=11, column=3).value = project.start_time
    worksheet.cell(row=12, column=1).value = 12
    worksheet.cell(row=12, column=2).value = "Ориентировочная дата оканчания проекта"
    worksheet.cell(row=12, column=3).value = project.finish_time
    worksheet.cell(row=13, column=1).value = 13
    worksheet.cell(row=13, column=2).value = "Ожидаемые результаты от реализации проекта"
    worksheet.cell(row=13, column=3).value = project.expecting_results
    worksheet.cell(row=14, column=1).value = 14
    worksheet.cell(row=14, column=2).value = "Создан в"
    worksheet.cell(row=14, column=3).value = str(project.created_by)
    workbook.save(response)

    return response


def export_excel_plan_graph_project(request, pk):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="plan_graph.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write header row
    header = ["ЭТАП", "НАИМЕНОВАНИЕ", "ДАТА НАЧАЛО", "ДАТА ОКОНЧАНИЕ", "СРОКИ В ДНЯХ", "СРОКИ В ЧАСАХ", "СТОИМОСТЬ"]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    project = ProjectPlan.objects.get(id=pk)
    stage_list = project.stage_list()

    counter = 2
    for stage in stage_list:
        worksheet.cell(row=counter, column=1).value = stage.stage_number
        worksheet.cell(row=counter, column=2).value = stage.description
        worksheet.cell(row=counter, column=3).value = stage.start_time
        worksheet.cell(row=counter, column=4).value = stage.finish_time
        worksheet.cell(row=counter, column=5).value = stage.duration_per_day
        worksheet.cell(row=counter, column=6).value = stage.duration_per_hour
        worksheet.cell(row=counter, column=7).value = stage.total_price_stage_and_task
        for task in stage.task_list():
            counter += 1
            worksheet.cell(row=counter, column=1).value = ""
            worksheet.cell(row=counter, column=2).value = task.description
            worksheet.cell(row=counter, column=3).value = task.start_time
            worksheet.cell(row=counter, column=4).value = task.finish_time
            worksheet.cell(row=counter, column=5).value = task.duration_per_day
            worksheet.cell(row=counter, column=6).value = task.duration_per_hour
            worksheet.cell(row=counter, column=7).value = task.total_price
        counter += 1
    project = ProjectPlan.objects.get(id=pk)
    worksheet.title = str(project.id)

    workbook.save(response)

    return response


def export_excel_cost_estimation_by_stage(request, pk):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="cost_estimation.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    project = ProjectPlan.objects.get(id=pk)

    # Write header row

    generation_excel_project_plan_table(project, worksheet)

    stage_list = project.stage_list()
    #
    column = 4
    for count, stage in enumerate(stage_list, start=1):
        worksheet.cell(row=1, column=column).value = f"Этап {count}"
        worksheet.cell(row=2, column=column).value = f"{stage.cost_price:,.2f}"
        worksheet.cell(row=3, column=column).value = f"{stage.salary_cost:,.2f}"
        worksheet.cell(row=4, column=column).value = f"{stage.period_expenses:,.2f}"
        worksheet.cell(row=5, column=column).value = f"{stage.contributions_to_IT_park:,.2f}"
        worksheet.cell(row=6, column=column).value = f"{stage.margin:,.2f}"
        worksheet.cell(row=7, column=column).value = f"{stage.total_price_stage_and_task:,.2f}"
        worksheet.column_dimensions[get_column_letter(column)].auto_size = True
        column += 1
    worksheet.title = "project"
    workbook.save(response)

    return response


def generation_excel_project_plan_table(project, worksheet):
    header = ["СТАТЬЯ ", "НОРМА, В %", "ВСЕГО ПО ПРОЕКТУ"]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 20
    worksheet.cell(row=2, column=1).value = "Себестоимость реализованной услуги"
    worksheet.cell(row=2, column=2).value = project.percent_cost_price
    worksheet.cell(row=2, column=3).value = f"{project.cost_price:,.2f}"
    worksheet.cell(row=3, column=1).value = "Себестоимость (Расходы на опл.труда)"
    worksheet.cell(row=3, column=2).value = project.percent_salary_cost
    worksheet.cell(row=3, column=3).value = f"{project.salary_cost:,.2f}"
    worksheet.cell(row=4, column=1).value = "Косвенные производственные затраты"
    worksheet.cell(row=4, column=2).value = project.percent_period_expenses
    worksheet.cell(row=4, column=3).value = f"{project.period_expenses:,.2f}"
    worksheet.cell(row=5, column=1).value = "Расходы на взносы IT-парк"
    worksheet.cell(row=5, column=2).value = 1
    worksheet.cell(row=5, column=3).value = f"{project.contributions_to_IT_park:,.2f}"
    worksheet.cell(row=6, column=1).value = "Чистый прибыль"
    worksheet.cell(row=6, column=2).value = project.percent_margin
    worksheet.cell(row=6, column=3).value = f"{project.margin:,.2f}"
    worksheet.cell(row=7, column=1).value = "Общая cтоимость проекта"
    worksheet.cell(row=7, column=2).value = ""
    worksheet.cell(row=7, column=3).value = f"{project.total_price_with_margin:,.2f}"


def export_excel_overall_project(request, pk):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="cost_estimation.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    project = ProjectPlan.objects.get(id=pk)
    worksheet.cell(row=1, column=1).value = "Количество этап:"
    worksheet.cell(row=1, column=2).value = project.stage_counter()
    worksheet.cell(row=2, column=1).value = "Ориентированные сроки начала :"
    worksheet.cell(row=2, column=2).value = project.start_time
    worksheet.cell(row=3, column=1).value = "Ориентированные сроки окончания проекта :"
    worksheet.cell(row=3, column=2).value = project.finish_time
    worksheet.cell(row=4, column=1).value = "Длителность проекта в днях:"
    worksheet.cell(row=4, column=2).value = project.duration_per_day
    worksheet.cell(row=5, column=1).value = "Длителность проекта в часах:"
    worksheet.cell(row=5, column=2).value = project.duration_per_hour

    header = ["СТАТЬЯ ", "НОРМА, В %", "ВСЕГО ПО ПРОЕКТУ"]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=7, column=col_num)
        cell.value = column_title
    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 20
    worksheet.cell(row=8, column=1).value = "Себестоимость реализованной услуги"
    worksheet.cell(row=8, column=2).value = project.percent_cost_price
    worksheet.cell(row=8, column=3).value = f"{project.cost_price:,.2f}"
    worksheet.cell(row=9, column=1).value = "Себестоимость (Расходы на опл.труда)"
    worksheet.cell(row=9, column=2).value = project.percent_salary_cost
    worksheet.cell(row=9, column=3).value = f"{project.salary_cost:,.2f}"
    worksheet.cell(row=10, column=1).value = "Косвенные производственные затраты"
    worksheet.cell(row=10, column=2).value = project.percent_period_expenses
    worksheet.cell(row=10, column=3).value = f"{project.period_expenses:,.2f}"
    worksheet.cell(row=11, column=1).value = "Расходы на взносы IT-парк"
    worksheet.cell(row=11, column=2).value = 1
    worksheet.cell(row=11, column=3).value = f"{project.contributions_to_IT_park:,.2f}"
    worksheet.cell(row=12, column=1).value = "Чистый прибыль"
    worksheet.cell(row=12, column=2).value = project.percent_margin
    worksheet.cell(row=12, column=3).value = f"{project.margin:,.2f}"
    worksheet.cell(row=13, column=1).value = "Общая cтоимость проекта"
    worksheet.cell(row=13, column=2).value = ""
    worksheet.cell(row=13, column=3).value = f"{project.total_price_with_margin:,.2f}"

    header = ["СТАТЬЯ ЗАТРАТ ", "КОММЕНТАРИЯ", "СУММА"]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=15, column=col_num)
        cell.value = column_title

    add_cost_list = project.additional_cost_plan_list()
    counter = 0

    for count, add_cost in enumerate(add_cost_list, start=16):
        worksheet.cell(row=count, column=1).value = add_cost.get_cost_type_display()
        worksheet.cell(row=count, column=2).value = add_cost.comment
        worksheet.cell(row=count, column=3).value = f"{add_cost.amount:,.2f}"
        counter = count

    counter += 2

    worksheet.cell(row=counter, column=1).value = "Общая cтоимость проекта+ Дополнительные расходы"
    worksheet.cell(row=counter, column=3).value = f"{project.total_price_with_additional_cost:,.2f}"

    worksheet.title = "project"
    workbook.save(response)

    return response


def export_excel_compare_plan_fact(request, pk):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="compare_plan_fact.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "test"

    header = [
        "№",
        "НАИМЕНОВАНИЕ",
        "Ответственный",
        "Департамент",
        "Выполнимость проекта",
        "Дедлайн",
        "Статус",
        "Проектная стоимость",
        "Фактическая стоимость",
        "Ставнение",
    ]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    project_fact_list = ProjectFact.objects.filter(deleted_status=False)
    user_role = request.user.user_role

    if user_role == UserRoleTypes.SUPER_USER or user_role == UserRoleTypes.FINANCE:
        excel_generation_plan_fact_compare(project_fact_list, worksheet)
    else:
        project_fact_list = project_fact_list.filter(created_by=request.user)
        excel_generation_plan_fact_compare(project_fact_list, worksheet)

    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col, width=20)
    worksheet.column_dimensions = dim_holder
    workbook.save(response)
    return response
